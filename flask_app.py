from flask import Flask, request, render_template, send_file
import openpyxl
from xlsx2html import xlsx2html
from weasyprint import HTML
import os
app = Flask(__name__)


@app.route('/')
def form():
    return render_template('index.html')


@app.route('/guardar_datos', methods=['POST'])
def submit():
    datos = request.get_json()  # Obtiene los datos JSON enviados desde el cliente
    wb = openpyxl.load_workbook('plantillaprueba.xlsx')
    ws = wb.active
    print(datos)
    ws['C7'] = datos['fecha']  # Fecha
    ws['I7'] = datos['hora']  # Hora
    ws['C8'] = datos['supervisor']  # Nombre Supervisor
    ws['C10'] = datos['nombreGranja']  # Nombre de la granja
    ws['B11'] = datos['direccion']  # Dirección de la granja
    ws['B12'] = datos['propietario']  # Propietario de la granja

    # Continúa escribiendo datos en las celdas correspondientes según el JSON recibido
    ws['C13'] = datos['tipoPollo']
    ws['B14'] = datos['destino']
    ws['C15'] = datos['cuadrilla1']
    ws['D15'] = datos['cuadrilla2']
    ws['F15'] = datos['cuadrilla3']
    ws['E15'] = datos['cuadrilla4']
    ws['C16'] = datos['cantOperarios1']
    ws['D16'] = datos['cantOperarios2']
    ws['F16'] = datos['cantOperarios3']
    ws['E18'] = datos['cantOperarios4']
    ws['B17'] = datos['mortandad']

    # Información sobre las personas presentes
    ws['E18'] = datos['propietarioPresente']
    ws['G18'] = datos['encargadoPresente']
    ws['I18'] = datos['ningunoPresente']

    # Estado de los caminos
    ws['E20'] = datos['tierra']
    ws['H20'] = datos['asfalto']
    ws['K20'] = datos['mejorado']

    # Datos sobre condiciones de caminos internos
    ws['G23'] = datos['caminosInternosB']
    ws['H23'] = datos['caminosInternosR']
    ws['J23'] = datos['caminosInternosM']
    ws['L23'] = datos['caminosInternosObservaciones']

    # Más datos sobre caminos
    ws['G24'] = datos['caminosHastaB']
    ws['H24'] = datos['caminosHastaR']
    ws['J24'] = datos['caminosHastaM']
    ws['L24'] = datos['caminosHastaObservaciones']

    # Datos sobre galpones y jaulas
    ws['G25'] = datos['tejidosGalponB']
    ws['H25'] = datos['tejidosGalponR']
    ws['J25'] = datos['tejidosGalponM']
    ws['L25'] = datos['tejidosGalponObservaciones']
    ws['G26'] = datos['camaGalponB']
    ws['H26'] = datos['camaGalponR']
    ws['J26'] = datos['camaGalponM']
    ws['L26'] = datos['camaGalponObservaciones']
    ws['G27'] = datos['estadoJaulasB']
    ws['H27'] = datos['estadoJaulasR']
    ws['J27'] = datos['estadoJaulasM']
    ws['L27'] = datos['estadoJaulasObservaciones']

    # Horarios y detalles adicionales
    ws['I30'] = datos['comienzoCarga']
    ws['B30'] = datos['corteAlimento']
    ws['B31'] = datos['horaAyuno']

    # Información sobre el equipo de carga
    ws['B35'] = datos['equipoCarga']

    # Observaciones específicas en áreas clave
    ws['G37'] = datos['capacitadoB']
    ws['H37'] = datos['capacitadoR']
    ws['I37'] = datos['capacitadoM']
    ws['J37'] = datos['capacitadoObservaciones']

    ws['G38'] = datos['manipulacionCargaB']
    ws['H38'] = datos['manipulacionCargaR']
    ws['I38'] = datos['manipulacionCargaM']
    ws['J38'] = datos['manipulacionCargaObservaciones']

    ws['G39'] = datos['encerradoAvesB']
    ws['H39'] = datos['encerradoAvesR']
    ws['I39'] = datos['encerradoAvesM']
    ws['J39'] = datos['encerradoAvesObservaciones']

    ws['G40'] = datos['cargaJaulasB']
    ws['H40'] = datos['cargaJaulasR']
    ws['I40'] = datos['cargaJaulasM']
    ws['J40'] = datos['cargaJaulasObservaciones']

    ws['G41'] = datos['avesMuertasRetiradasB']
    ws['H41'] = datos['avesMuertasRetiradasR']
    ws['I41'] = datos['avesMuertasRetiradasM']
    ws['J41'] = datos['avesMuertasRetiradasObservaciones']

    # Indicadores de bienestar animal
    ws['G44'] = datos['cargaAvesMuertasSi']
    ws['H44'] = datos['cargaAvesMuertasNo']
    ws['I44'] = datos['cargaAvesMuertasNa']
    ws['G45'] = datos['sacrificioBienestarSi']
    ws['H45'] = datos['sacrificioBienestarNo']
    ws['I45'] = datos['sacrificioBienestarNa']

    # Otros campos
    ws['B52'] = datos['na']
    ws['B54'] = datos['observaciones']

    formularioxlsx = 'CPLF00-00 Planilla de Inspección de Granjas - Carga de Aves.xlsx'
    wb.save(formularioxlsx)

    xlsx_file = formularioxlsx
    html_file = 'formulario.html'
    xlsx2html(xlsx_file, html_file)

    

    # Convertir el HTML a PDF
    pdf_file = 'formulario.pdf'

    HTML(html_file).write_pdf(
        pdf_file, stylesheets=[os.path.join('styles.css')])

    return send_file(pdf_file, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
